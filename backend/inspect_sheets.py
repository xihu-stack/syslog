"""只看结构：打印每个 sheet 的第1行(标题)和第2行(表头)，不打印任何数据行/个人信息。"""
import openpyxl

FILES = [
    r"C:\Users\huxi\Desktop\111.xlsx",
    r"C:\Users\huxi\Desktop\222.xlsx",
]

for path in FILES:
    print("=" * 70)
    print("文件:", path.split("\\")[-1])
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n  [{ws.title}]  行={ws.max_row} 列={ws.max_column}")
        for r in (1, 2, 3):
            if r <= ws.max_row:
                vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                print(f"   第{r}行:", vals)
