"""IP-Guard Excel 导出解析器：文档操作 / 网页浏览 / 关键字搜索 三类日志。

按 sheet 名分发；自动识别表头行（不写死行号）；Excel 日期序列→datetime；
文件大小→字节；操作类型/磁盘类型/域名分类归一化。
"""
from __future__ import annotations

from datetime import datetime, timedelta
from urllib.parse import urlparse

from models import CanonicalEvent, Category, DOC_ACTION_MAP
import dicts

# IP-Guard 计算机名前缀模式（如 HLX-BJ- / HLX-SZ-），提取纯姓名用于跨源合并
import re as _re
def _extract_name(computer):
    """从 IP-Guard 计算机名提取纯姓名：HLX-BJ-孙翔宇 → 孙翔宇"""
    m = _re.match(r'^[A-Z]+-[A-Z]+-(.+)$', computer or '')
    return m.group(1) if m else (computer or '')


# ---------------- 通用工具 ----------------

def excel_serial_to_dt(serial) -> datetime:
    if isinstance(serial, datetime):
        return serial
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(serial))
    except (TypeError, ValueError):
        return datetime.min


_SIZE_UNITS = {"B": 1, "BYTES": 1, "KB": 1024, "MB": 1024 ** 2, "GB": 1024 ** 3, "TB": 1024 ** 4}


def parse_size(text) -> int:
    if text is None:
        return 0
    s = str(text).strip().replace(",", "")
    if not s or s == "0":
        return 0
    parts = s.split()
    try:
        if len(parts) == 2:
            return int(float(parts[0]) * _SIZE_UNITS.get(parts[1].upper(), 1))
        return int(float(parts[0]))
    except (ValueError, IndexError):
        return 0


def normalize_user(raw_user) -> str:
    if raw_user is None:
        return ""
    u = str(raw_user).strip()
    if "\\" in u:
        u = u.rsplit("\\", 1)[-1]
    return u


def disk_to_channel(disk_type: str) -> str:
    d = (disk_type or "").strip()
    if not d or d in ("硬盘", "本地盘", "本地硬盘"):
        return "LOCAL"
    if "移动" in d or "U盘" in d or "可移动" in d or "USB" in d.upper():
        return "USB"
    if "网络" in d or "网盘" in d or "共享" in d:
        return "NETWORK"
    return "OTHER"


# ---------------- 域名分类（外发通道 / 招聘，字典可配置）----------------

def url_domain(url: str) -> str:
    try:
        netloc = urlparse(url if "://" in url else "http://" + (url or "")).netloc.lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ""


def domain_class(domain: str) -> str:
    d = (domain or "").lower()
    if any(x in d for x in dicts.get("netdisk_domains")):
        return "netdisk"
    if any(x in d for x in dicts.get("personal_email_domains")):
        return "personal_email"
    if any(x in d for x in dicts.get("recruitment_sites")):
        return "recruitment"
    return "other"


# ---------------- 表头识别 ----------------

HEADER_ALIASES = {
    "序号": "no", "类型": "type", "时间": "time",
    "计算机": "computer", "计算机组": "computer_group",
    "用户": "user", "用户组": "user_group",
    "源文件": "source_file", "文件名": "source_file",
    "文件大小": "size", "大小": "size",
    "路径": "path", "磁盘类型": "disk_type",
    "应用程序": "application", "进程": "application",
    "标题": "title", "窗口标题": "title",
    "网址": "url", "url": "url", "域名": "search_engine",
    "搜索关键字": "keyword", "关键字": "keyword", "搜索词": "keyword",
}


def find_header(ws, required: tuple[str, ...], max_scan: int = 6) -> tuple[int, dict]:
    """在前 max_scan 行找同时含 required 字段的表头行。返回 (行号, {字段:列索引})。"""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        mapping: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value).strip()
            if val in HEADER_ALIASES:
                mapping[HEADER_ALIASES[val]] = c
        if all(k in mapping for k in required):
            return r, mapping
    return -1, {}


# ---------------- 文档操作日志 ----------------

def parse_doc_log_sheet(ws) -> list[CanonicalEvent]:
    header_row, col = find_header(ws, ("type", "time"))
    if header_row < 0:
        return []
    events: list[CanonicalEvent] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def g(key):
            return ws.cell(r, col[key]).value if key in col else None
        op_type = str(g("type") or "").strip()
        if not op_type or op_type == "None":
            continue
        computer = str(g("computer") or "").strip()
        disk = str(g("disk_type") or "").strip()
        events.append(CanonicalEvent(
            occurred_at=excel_serial_to_dt(g("time")),
            employee_id=_extract_name(computer),  # 提取纯姓名，跨源合并
            source="ipguard",
            device_id=computer, category=Category.DOC.value,
            action=DOC_ACTION_MAP.get(op_type, "UNKNOWN"), target_type="FILE",
            target_value=str(g("source_file") or "").strip(),
            size_bytes=parse_size(g("size")), count=1,
            raw={"op_type": op_type, "computer_group": str(g("computer_group") or "").strip(),
                 "user_group": str(g("user_group") or "").strip(), "user_raw": str(g("user") or "").strip(),
                 "path": str(g("path") or "").strip(), "disk_type": disk,
                 "channel": disk_to_channel(disk),
                 "application": str(g("application") or "").strip(),
                 "title": str(g("title") or "").strip()},
        ))
    return events


# ---------------- 网页浏览日志 ----------------

def parse_web_log_sheet(ws) -> list[CanonicalEvent]:
    header_row, col = find_header(ws, ("time", "url"))
    if header_row < 0:
        return []
    events: list[CanonicalEvent] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def g(key):
            return ws.cell(r, col[key]).value if key in col else None
        url = str(g("url") or "").strip()
        if not url or url == "None":
            continue
        computer = str(g("computer") or "").strip()
        domain = url_domain(url)
        events.append(CanonicalEvent(
            occurred_at=excel_serial_to_dt(g("time")),
            employee_id=_extract_name(computer),  # 提取纯姓名，跨源合并
            source="ipguard",
            device_id=computer, category=Category.WEB.value,
            action="VISIT", target_type="URL", target_value=url,
            raw={"title": str(g("title") or "").strip(), "url": url, "domain": domain,
                 "domain_class": domain_class(domain),
                 "computer_group": str(g("computer_group") or "").strip(),
                 "user_group": str(g("user_group") or "").strip(),
                 "user_raw": str(g("user") or "").strip()},
        ))
    return events


# ---------------- 关键字搜索日志 ----------------

def parse_search_log_sheet(ws) -> list[CanonicalEvent]:
    header_row, col = find_header(ws, ("time", "keyword"))
    if header_row < 0:
        return []
    events: list[CanonicalEvent] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def g(key):
            return ws.cell(r, col[key]).value if key in col else None
        kw = str(g("keyword") or "").strip()
        if not kw or kw == "None":
            continue
        computer = str(g("computer") or "").strip()
        events.append(CanonicalEvent(
            occurred_at=excel_serial_to_dt(g("time")),
            employee_id=_extract_name(computer),  # 提取纯姓名，跨源合并
            source="ipguard",
            device_id=computer, category=Category.SEARCH.value,
            action="SEARCH", target_type="URL", target_value=kw,
            raw={"keyword": kw,
                 "search_engine": str(g("search_engine") or "").strip(),
                 "application": str(g("application") or "").strip(),
                 "computer_group": str(g("computer_group") or "").strip(),
                 "user_group": str(g("user_group") or "").strip(),
                 "user_raw": str(g("user") or "").strip()},
        ))
    return events


# ---------------- 入口：按 sheet 名分发 ----------------

def parse_ipguard_excel(path: str) -> list[CanonicalEvent]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    events: list[CanonicalEvent] = []
    for ws in wb.worksheets:
        name = (ws.title or "").strip()
        if "文档" in name:
            events.extend(parse_doc_log_sheet(ws))
        elif "网页浏览" in name or "上网" in name or "网页" in name:
            events.extend(parse_web_log_sheet(ws))
        elif "搜索" in name or "关键字" in name:
            events.extend(parse_search_log_sheet(ws))
        else:
            # 兜底：按表头特征尝试
            ev = parse_doc_log_sheet(ws) + parse_web_log_sheet(ws) + parse_search_log_sheet(ws)
            events.extend(ev)
    return events


if __name__ == "__main__":
    import sys
    from run_parse import main
    sys.argv = ["run_parse.py", sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\huxi\Desktop\111.xlsx"]
    main()
