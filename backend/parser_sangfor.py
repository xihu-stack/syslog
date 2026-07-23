"""深信服上网行为管理日志解析器（导出 CSV/xlsx，表头驱动）。

按表头列名识别（序号/用户名/网站分类/访问域名/URL地址/时间…），兼容列顺序与缺列。
输出标准 WEB 事件，raw 里带深信服的 网站分类（高质量信号，喂 AI 用）。
"""
from __future__ import annotations

import csv
import os
import re
from datetime import datetime

from models import CanonicalEvent
from parser_ipguard import url_domain


def _clean_user(u):
    """清理用户名：周珈妍(周珈妍) → 周珈妍；去掉尾部括号。"""
    u = (u or "").strip()
    u = re.sub(r'^(.+?)\(\1\)$', r'\1', u)  # 名字(同名) → 名字
    u = re.sub(r'\([^)]*\)$', '', u).strip()  # 去尾部括号
    return u

ALIASES = {
    "序号": "no", "用户名": "user", "组名": "group", "源IP": "src_ip", "终端类型": "terminal",
    "位置": "location", "目标IP": "dst_ip", "网站分类": "category", "标题": "title",
    "访问域名": "domain", "URL地址": "url", "访问控制": "access", "时间": "time",
    "解密情况": "decrypt", "详情": "detail",
}


def _map_headers(cols):
    m = {}
    for i, c in enumerate(cols):
        c = str(c or "").strip()
        if c in ALIASES:
            m[ALIASES[c]] = i
    return m


def _parse_time(s):
    s = str(s or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return datetime.min


def _emit(row, hdr):
    def g(k):
        return (row[hdr[k]] if k in hdr and hdr[k] < len(row) else "").strip()
    domain = g("domain")
    if not domain:
        return None
    user = _clean_user(g("user"))
    url = g("url") or domain
    return CanonicalEvent(
        occurred_at=_parse_time(g("time")),
        employee_id=user or g("src_ip") or "unknown",
        device_id=g("src_ip"),
        category="WEB", action="VISIT", target_type="URL", target_value=url, count=1,
        raw={"domain": domain, "category": g("category"), "group": g("group"),
             "terminal": g("terminal"), "src_ip": g("src_ip"), "dst_ip": g("dst_ip"),
             "title": g("title"), "access": g("access"), "decrypt": g("decrypt")})


def _parse_rows(rows):
    hdr, start = None, 0
    for i, r in enumerate(rows):
        m = _map_headers(r)
        if "user" in m and "time" in m and "domain" in m:
            hdr, start = m, i + 1
            break
    if not hdr:
        return []
    events = []
    for r in rows[start:]:
        if not r or len(r) < 3:
            continue
        ev = _emit(r, hdr)
        if ev:
            events.append(ev)
    return events


SANGFOR_SYSLOG_FIELDS = ["record_time", "user", "group", "host_ip", "dst_ip", "serv",
                         "app", "site", "tm_type", "net_action", "url", "DNS", "title", "snapshot"]


def parse_sangfor_syslog(raw_msg: str):
    """解析深信服 syslog 报文 → CanonicalEvent。支持 [key:value]、key=value、tab 三种格式。"""
    msg = raw_msg.strip()
    # 去掉 syslog 头部 <PRI>timestamp host tag:
    msg = re.sub(r'^<\d+>', '', msg).strip()
    msg = re.sub(r'^\w{3}\s+\d+\s+\d+:\d+:\d+\s+\S+\s+\S+:\s*', '', msg).strip()

    fields = {}
    if '[' in msg and ']' in msg:
        # 深信服标准 [key:value] 格式
        for m in re.finditer(r'\[(\w+):([^\]]*)\]', msg):
            fields[m.group(1)] = m.group(2).strip()
    elif "=" in msg and ("record_time=" in msg or "user=" in msg):
        for m in re.finditer(r'(\w+)=(.*?)(?=\s+\w+=|$)', msg):
            fields[m.group(1)] = m.group(2).strip().strip('"')
    else:
        parts = msg.split("\t") if "\t" in msg else msg.split()
        for i, fn in enumerate(SANGFOR_SYSLOG_FIELDS):
            if i < len(parts):
                fields[fn] = parts[i]

    user = _clean_user(fields.get("user", ""))
    url = (fields.get("url") or "").strip()
    if not user or not url or url == "-":
        return None
    domain = url_domain(url)
    # syslog 里 app=网站分类(IT行业)，site=位置(未定义位置)
    category = fields.get("app", "") or fields.get("site", "")
    return CanonicalEvent(
        occurred_at=_parse_time(fields.get("record_time", "")),
        employee_id=user or fields.get("host_ip", ""),
        device_id=fields.get("host_ip", ""),
        category="WEB", action="VISIT", target_type="URL", target_value=url, count=1,
        raw={"domain": domain, "category": category, "group": fields.get("group", ""),
             "src_ip": fields.get("host_ip", ""), "dst_ip": fields.get("dst_ip", ""),
             "app": fields.get("app", ""), "serv": fields.get("serv", ""),
             "title": fields.get("title", ""), "net_action": fields.get("net_action", "")})


def parse_sangfor(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            sample = f.readline()
            f.seek(0)
            delim = "\t" if sample.count("\t") > sample.count(",") else ","
            return _parse_rows(list(csv.reader(f, delimiter=delim)))
    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        events = []
        for ws in wb.worksheets:
            rows = [[(c.value if c.value is not None else "") for c in row]
                    for row in ws.iter_rows(values_only=False)]
            # 取每行每列的值
            data = [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]
            events.extend(_parse_rows(data))
        return events
    return []
